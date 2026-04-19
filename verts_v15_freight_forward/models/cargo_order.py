# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

from datetime import datetime, timedelta
from odoo import api, fields, models, SUPERUSER_ID, _
from odoo.exceptions import AccessError, UserError, ValidationError
from odoo.tools.misc import formatLang, get_lang
from odoo.osv import expression
from odoo.tools import float_is_zero, float_compare
from num2words import num2words
import xlwt
import xlrd
import base64
import lxml
from lxml import etree
from io import BytesIO
import os
import openpyxl
import io
import xlsxwriter
# from xlutils.copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font
import tempfile
from dataclasses import field


class CargoOrder(models.Model):
    _name = "cargo.order"
    _inherit = ['portal.mixin', 'mail.thread', 'mail.activity.mixin', 'utm.mixin']
    _description = "Cargo Order"
    _order = 'date_order desc, id desc'
    _check_company_auto = True

    def get_day_suffix(self, day):
        if 11 <= day <= 13:
            return 'th'
        else:
            return {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')

    def action_download_excel(self):
        """ Method to modify and download an Excel file stored in the module folder """
        module_path = os.path.dirname(os.path.abspath(__file__))  # Get module path
        file_path = os.path.join(module_path, '../static/excel/AWB.xlsx')

        wb = load_workbook(file_path)
        ws = wb.active

        airline_no = self.shipping_line and self.shipping_line.airline_no or ''
        small_font = Font(size=5)  # Adjust size as needed

        origin_port_code = ''
        if self.port_of_loading_id and len(self.port_of_loading_id.name) >= 3:
            origin_port_code = self.port_of_loading_id.name[-3:]

        awb_no = ''
        # if self.mawb and len(self.mawb.awb_no) >= 8:
        #     awb_no = self.mawb.awb_no[-8:]
        if self.mawb and self.mawb.awb_no:
            awb_no = self.mawb.awb_no

        if airline_no and len(airline_no) >2:
            ws["A1"] = airline_no[0] or ''
            ws["B1"] = airline_no[1] or ''
            ws["C1"] = airline_no[2] or ''

        ws["D1"] = origin_port_code or ''
        ws["G1"] = awb_no or ''
        ws["AK1"] = airline_no+'-'+awb_no
        ws["AH58"] = origin_port_code or ''
        ws["AC61"] = airline_no + '-' + awb_no

        current_date = datetime.now().strftime("%d-%m-%Y")
        ws["S58"] = current_date or ''
        ws["S54"] = self.sign_of_shipper or ''
        ws["AN58"] = self.sign_of_carrier or ''

        # ws["A3"] = self.consignor_id.name or ''
        consignor_address = ''
        if self.consignor_id.name:
            consignor_address += f"{self.consignor_id.name},\n"
        if self.consignor_id.street:
            consignor_address += f"{self.consignor_id.street}, \n"
        if self.consignor_id.street2:
            consignor_address += f"{self.consignor_id.street2}, \n"
        if self.consignor_id.city_id:
            consignor_address += f"{self.consignor_id.city_id.name}, "
        if self.consignor_id.state_id:
            consignor_address += f"{self.consignor_id.state_id.name}, "
        if self.consignor_id.country_id:
            consignor_address += f"{self.consignor_id.country_id.name}, "
        if self.consignor_id.zip:
            consignor_address += f"{self.consignor_id.zip}, "
        if self.consignor_id.mobile:
            consignor_address += f"\nMob: {self.consignor_id.mobile}, "
        if self.consignor_id.email:
            consignor_address += f"\n{self.consignor_id.email}, "
        ws["A4"] = consignor_address

        # Consignee name and address
        # ws["A9"] = self.consignee_id.name or ''
        consignee_address = ''
        if self.consignee_id.name:
            consignee_address += f"{self.consignee_id.name},\n"
        if self.consignee_id.street:
            consignee_address += f"{self.consignee_id.street}, \n"
        if self.consignee_id.street2:
            consignee_address += f"{self.consignee_id.street2}, \n"
        if self.consignee_id.city_id:
            consignee_address += f"{self.consignee_id.city_id.name}, "
        if self.consignee_id.state_id:
            consignee_address += f"{self.consignee_id.state_id.name}, "
        if self.consignee_id.country_id:
            consignee_address += f"{self.consignee_id.country_id.name}, "
        if self.consignee_id.zip:
            consignee_address += f"{self.consignee_id.zip}, "
        if self.consignee_id.mobile:
            consignee_address += f"\nMob: {self.consignee_id.mobile}, "
        if self.consignee_id.email:
            consignee_address += f"\n{self.consignee_id.email}, "
        ws["A11"] = consignee_address

        if self.agent_id.city_id:
            ws["A18"] = f"{self.agent_id.name or ''}, {self.agent_id.city_id.name or ''}"
        else:
            ws["A18"] = self.agent_id.name or ''

        ws["A23"] = self.port_of_loading_id.name or ''
        ws["A28"] = self.port_of_discharge_id.name or ''
        # ws["H29"] = self.dept_date or ''
        ws["M3"] = self.consignee_account or ''
        ws["M10"] = self.consignor_account or ''
        # ws["L29"] = self.dept_date or ''
        ws["X6"] = self.shipping_line.name or ''
        ws["X26"] = self.currency_id.name or ''
        ws["A30"] = self.handling_info or ''
        ws["X18"] = self.accounting_info or ''
        # ws["V27"] = self.other_coll or ''
        ws["AK26"] = self.declare_value or ''
        ws["AR26"] = self.declare_value_custom or ''
        ws["X28"] = self.amount_insurance or ''
        ws["C44"] = self.total_gross_weight_cargo or ''
        ws["AD44"] = self.total_amount_cargo or ''
        # ws["J54"] = self.total_chargeable_weight_cargo or ''
        ws["A21"] = self.agent_code or ''
        ws["L21"] = self.agent_account_number or ''
        # ws["A27"] = self.to_1.name or ''
        # ws["J27"] = self.to_2.name or ''
        # ws["M27"] = self.to_3.name or ''
        ws["C26"] = self.airline_code_1 or ''
        ws["Q26"] = self.airline_code_2 or ''
        ws["V26"] = self.airline_code_3 or ''

        ws["AG23"] = self.optional_shipping or ''
        ws["X23"] = self.ref_num or ''
        ws["AM44"] = self.other_remarks or ''

        combined_value1 =''
        combined_value2 = ''
        if self.flight_date_1:
            day = self.flight_date_1.day
            suffix = self.get_day_suffix(day)
            formatted_date = f"{day}{suffix} {self.flight_date_1.strftime('%B')}"  # e.g., 7th May
            combined_value1 = f"{self.flight_number_1}/{formatted_date}" if self.flight_number_1 else formatted_date
        else:
            combined_value1 = self.flight_number_1 or ''

        if self.flight_date_2:
            day = self.flight_date_2.day
            suffix = self.get_day_suffix(day)
            formatted_date = f"{day}{suffix} {self.flight_date_2.strftime('%B')}"  # e.g., 7th May
            combined_value2 = f"{self.flight_number_2}/{formatted_date}" if self.flight_number_2 else formatted_date
        else:
            combined_value2 = self.flight_number_2 or ''

        ws["M28"] = combined_value1
        ws["M28"].font = small_font

        ws["R28"] = combined_value2
        ws["R28"].font = small_font

        if self.to_1 and len(self.to_1.name) >= 3:
            ws["A26"] = self.to_1.name[-3:]

        if self.to_2 and len(self.to_2.name) >= 3:
            ws["O26"] = self.to_2.name[-3:]

        if self.to_3 and len(self.to_3.name) >= 3:
            ws["T26"] = self.to_3.name[-3:]

        nature_goods = ''
        if self.cargo_details_ids:
            total_rcp = 0.0
            line_no = 36

            for line in self.cargo_details_ids:
                ws[f"A{line_no}"] = line.no_of_pieces
                ws[f"C{line_no}"] = line.gross_weight
                ws[f"G{line_no}"] = line.uom_id.name or ''
                ws[f"I{line_no}"] = line.class_cargo or ''
                ws[f"J{line_no}"] = line.remarks or ''
                ws[f"O{line_no}"] = line.chargeable_weight
                ws[f"V{line_no}"] = line.rate
                ws[f"AD{line_no}"] = line.total

                # if line.remarks:
                #     nature_goods += f"{line.remarks}, "

                total_rcp += line.no_of_pieces
                line_no += 1

            ws["A44"] = total_rcp

        if self.nature_goods:
            # nature_goods += f"\n{self.nature_goods}, "
            ws["AM35"] = self.nature_goods or ''

        # if self.commodity:
        #     nature_goods += f"{self.commodity.name}, "

        if self.dimension_volume =='dimension' and self.opportunity_id:
            for line in self.opportunity_id.order_line:
                nature_goods += f"\n{line.length_per_package}×{line.width_per_package}×{line.height_per_package},"
        elif self.dimension_volume == 'volume':
            nature_goods += f"\nTotal Volume: {self.opportunity_id.sum_total_volume_cbm}, "

        if self.hs_code:
            nature_goods += f"\nHS Code-{self.hs_code}, "

        ws["AM37"] = nature_goods or ''

        filter_lines = self.order_line.filtered(lambda line: line.show_in_awb == True)
        # oth_charges = ''
        count =1
        for fl in filter_lines:
            if count == 1:
                ws["T46"] = fl.product_id.charge_code or ''
                ws["X46"] = fl.price_total or ''
            if count == 2:
                ws["AD46"] = fl.product_id.charge_code or ''
                ws["AH46"] = fl.price_total or ''
            if count == 3:
                ws["AN46"] = fl.product_id.charge_code or ''
                ws["AR46"] = fl.price_total or ''
            if count == 4:
                ws["T48"] = fl.product_id.charge_code or ''
                ws["X48"] = fl.price_total or ''
            if count == 5:
                ws["AD48"] = fl.product_id.charge_code or ''
                ws["AH48"] = fl.price_total or ''
            if count == 6:
                ws["AN48"] = fl.product_id.charge_code or ''
                ws["AR48"] = fl.price_total or ''
            count +=1

            # oth_charges += f"{fl.product_id.charge_code}-{fl.price_total},\n"
        # ws["T46"] = oth_charges or ''

        # agent_lines = self.order_line.filtered(lambda line: line.charge_type == 'due_agent')
        # due_agent_total = 0.0
        # for al in agent_lines:
        #     due_agent_total += al.price_total
        #
        # carrier_lines = self.order_line.filtered(lambda line: line.charge_type == 'due_carrier')
        # due_carrier_total = 0.0
        # for cl in carrier_lines:
        #     due_carrier_total += cl.price_total

        if self.code == 'pp':
            ws["AA26"] = "PP"
            ws["AC26"] = "X"  # ppd
            ws["AG26"] = "X"  # other_ppd
            ws["AE26"] = ""  # Clear coll
            ws["AI26"] = ""  # Clear other_coll
            ws["A52"] = self.due_agent
            ws["A54"] = self.due_carrier
            ws["A58"] = self.due_agent+self.total_amount_cargo+self.due_carrier
            ws["A46"] = self.total_amount_cargo

        elif self.code == 'cc':
            ws["AA26"] = "CC"
            ws["AE26"] = "X"  # coll
            ws["AI26"] = "X"  # other_coll
            ws["AC26"] = ""  # Clear ppd
            ws["AE26"] = ""  # Clear other_ppd
            ws["J52"] = self.due_agent
            ws["J54"] = self.due_carrier
            ws["J58"] = self.due_agent+self.total_amount_cargo+self.due_carrier
            ws["J46"] = self.total_amount_cargo


            # Apply font size 6 to all filled cells
        # small_font = Font(size=6)
        # for row in ws.iter_rows():
        #     for cell in row:
        #         if cell.value:  # Apply only to non-empty cells
        #             cell.font = small_font
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # Get the original font and update only the size
                    original_font = cell.font
                    cell.font = Font(
                        name=original_font.name,
                        size=original_font.size,
                        bold=original_font.bold,
                        italic=original_font.italic,
                        vertAlign=original_font.vertAlign,
                        underline=original_font.underline,
                        strike=original_font.strike,
                        color=original_font.color
                    )

        # Save the modified file to a temporary path
        # temp_path = os.path.join(module_path, '../static/excel/temp_AWB.xlsx')
        # wb.save(temp_path)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            temp_path = tmp.name
            wb.save(temp_path)

        # Convert to Base64
        with open(temp_path, 'rb') as file:
            file_data = file.read()
            file_base64 = base64.b64encode(file_data)

        filename = f'AWB_{self.name}.xlsx'
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': file_base64,
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary',
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }


    def _default_validity_date(self):
        if self.env['ir.config_parameter'].get_param('sale.use_quotation_validity_days'):
            days = self.env.company.quotation_validity_days
            if days > 0:
                return fields.Date.to_string(datetime.now() + timedelta(days))
        return False

    @api.depends('order_line.price_total')
    def _amount_all(self):
        """
        Compute the total amounts of the SO.
        """
        for order in self:
            amount_untaxed = amount_tax = 0.0
            for line in order.order_line:
                amount_untaxed += line.price_subtotal
                amount_tax += line.price_tax
            order.update({
                'amount_untaxed': amount_untaxed,
                'amount_tax': amount_tax,
                'amount_total': amount_untaxed + amount_tax,
            })

    @api.model
    def _default_note(self):
        return self.env['ir.config_parameter'].get_param(
            'account.use_invoice_terms') and self.env.company.invoice_terms or ''

    @api.model
    def _get_default_team(self):
        return self.env['crm.team']._get_default_team_id()

    @api.onchange('fiscal_position_id')
    def _compute_tax_id(self):
        """
        Trigger the recompute of the taxes if the fiscal position is changed on the SO.
        """
        for order in self:
            order.order_line._compute_tax_id()

    def _get_default_dimension_uom_id(self):
        return self.env.ref('uom.product_uom_cm')

    def _get_default_weight_uom_id(self):
        return self.env.ref('uom.product_uom_kgm')

    @api.depends('sale_expense_line', 'order_line')
    def get_total_expense(self):
        expenses_total = 0.0
        for sline in self.sale_expense_line:
            expenses_total += sline.exp_amt
        line_total_qty = 0.0
        for line in self.order_line:
            line_total_qty += line.product_uom_qty
        if line_total_qty > 0 and expenses_total > 0:
            self.total_expense = expenses_total
            self.rate_per_kg_on_expense = expenses_total / line_total_qty

    name = fields.Char(string='Order Reference', required=True, copy=False, readonly=True,
                       index=True, default=lambda self: _('New'))
    origin = fields.Char(string='Source Document',
                         help="Reference of the document that generated this sales order request.")
    client_order_ref = fields.Char(string='Customer Reference', copy=False)
    # profitability = fields.Float(string='Profitability',compute='_compute_profatibility')##by kajal
    reference = fields.Char(string='Payment Ref.', copy=False,
                            help='The payment communication of this sale order.')
    invoice_amount = fields.Char(string="invoice Amount")
    sale_id = fields.Many2one('sale.order', string="Sale Order", tracking=True)
    state = fields.Selection([
        ('cargo_ready', 'Cargo Ready'),
        ('cro_requested', 'CRO Requested'),
        ('cro_released', 'CRO Released'),
        ('stuff_started', 'Stuff Started'),
        ('stuffing_completed', 'Stuffing Completed'),
        ('sent_to_port', 'Sent to Port'),
        ('custom_clearance', 'Custom Clearance'),
        ('sealing_done', 'Sealing Done'),
        ('bill_of_lading_completed', 'Bill of Lading Completed'),
        ('vessel_set_to_sail', 'Vessel Set to Sail'),
        ('pre_alert', 'PreAlert'),
        ('destination_arrival', 'Destination Arrival'),
        ('destination_clearance', 'Destination Clearance'),
        ('delivery', 'Delivery'),
        ('completed', 'Completed'),
        ('book_with_the_flight', 'Book with the flight'),
        ('arrange_the_airway_bill', 'Arrange the Airway Bill'),
        ('pickup', 'Pickup'),
        ('booking_flight_schedule_confirmation', 'Booking/Flight Schedule Confirmation'),
        ('prealert_received', 'Prealert Received'),
        ('eta_shipment_arrival', 'ETA - Shipment Arrival'),
        ('document_readiness', 'Document Readiness'),
        ('cargo_arrival_notice', 'Cargo Arrival Notice'),
        ('forward_for_clearance', 'Forward for Clearance'),
        ('under_clearance', 'Under Clearance'),
        ('awaiting_approval', 'Awaiting Approval'),
        ('custom_declaration_completed', 'Custom Declaration Completed'),
        ('under_custom_inspection', 'Under Custom Inspection'),
        ('out_for_delivery', 'Out for Delivery'),
        ('draft', 'Draft'),
        ('request_confirm', 'Request Confirm'),
        ('picking', 'Picking'),
        ('packing', 'Packing'),
        ('material_received', 'Material Received'),
        ('material_received_done', 'Material Received Done')],
        string='Status', readonly=True, copy=False, index=True, tracking=3)

    date_order = fields.Datetime(string='Order Date', required=True, index=True, copy=False,
                                 default=fields.Datetime.now,
                                 help="Creation date of draft/sent orders,\nConfirmation date of confirmed orders.")
    validity_date = fields.Date(string='Expiration', copy=False, default=_default_validity_date)
    is_expired = fields.Boolean(compute='_compute_is_expired', string="Is expired")

    create_date = fields.Datetime(string='Creation Date', readonly=True, index=True,
                                  help="Date on which sales order is created.")
    # order type selection
    # internal_type = fields.Selection([
    #     ('sea_outbound', 'Sea Outbound'),
    #     ('sea_inbound', 'Sea Inbound'),
    #     ('air_outbound', 'Air Outbound'),
    #     ('air_inbound', 'Air Inbound'),
    #     ('custom_clearance', 'Custom Clearance'),
    #     # ('packing_and_removal', 'Packing And Removal'),
    #     # ('warehousing', 'Warehousing')
    # ], string='Internal Type')
    order_type = fields.Selection([
        ('freight', 'Freight'),
        ('custom', 'Custom'),
        # ('packing_and_removal', 'Packing And Removal'),
        # ('warehousing', 'Warehousing')
    ], string='Order Type')
    mode = fields.Selection([
        ('air', 'Air'),
        ('sea', 'Sea'),
        ('land', 'Land'),
        ('courier', 'Courier')
    ], string='Mode')
    import_export = fields.Selection([
        ('import', 'Import'),
        ('export', 'Export'),
        ('cross_trade', 'Cross Trade')
    ], string='Import/Export')
    user_id = fields.Many2one(
        'res.users', string='Salesperson', index=True, tracking=2, default=lambda self: self.env.user,
        domain=lambda self: [('groups_id', 'in', self.env.ref('sales_team.group_sale_salesman').id)])
    partner_id = fields.Many2one(
        'res.partner', string='Customer',
        required=True, change_default=True, index=True, tracking=1,
        domain="['|', ('company_id', '=', False), ('company_id', '=', company_id)]")
    partner_invoice_id = fields.Many2one(
        'res.partner', string='Invoice Address',
        domain="['|', ('company_id', '=', False), ('company_id', '=', company_id)]", )
    partner_shipping_id = fields.Many2one(
        'res.partner', string='Delivery Address',
        domain="['|', ('company_id', '=', False), ('company_id', '=', company_id)]", )

    pricelist_id = fields.Many2one(
        'product.pricelist', string='Pricelist', check_company=True,
        domain="['|', ('company_id', '=', False), ('company_id', '=', company_id)]", tracking=1,
        help="If you change the pricelist, only newly added lines will be affected.")
    currency_id = fields.Many2one(related='pricelist_id.currency_id', depends=["pricelist_id"], store=True, string='Currency')
    order_line = fields.One2many('cargo.order.line', 'order_id', string='Order Lines',
                                 states={'cancel': [('readonly', True)], 'done': [('readonly', True)]}, copy=True,
                                 auto_join=True)

    note = fields.Text('Terms and conditions', default=_default_note)

    amount_untaxed = fields.Monetary(string='Untaxed Amount', store=True, readonly=True, compute='_amount_all',
                                     tracking=5)
    amount_by_group = fields.Binary(string="Tax amount by group", compute='_amount_by_group',
                                    help="type: [(name, amount, base, formated amount, formated base)]")
    amount_tax = fields.Monetary(string='Taxes', store=True, readonly=True, compute='_amount_all')
    amount_total = fields.Monetary(string='Total', store=True, readonly=True, compute='_amount_all', tracking=4)
    currency_rate = fields.Float("Currency Rate", compute='_compute_currency_rate', compute_sudo=True, store=True,
                                 digits=(12, 6), readonly=True,
                                 help='The rate of the currency to the currency of rate 1 applicable at the date of the order')

    payment_term_id = fields.Many2one(
        'account.payment.term', string='Payment Terms', check_company=True,  # Unrequired company
        domain="['|', ('company_id', '=', False), ('company_id', '=', company_id)]", )
    fiscal_position_id = fields.Many2one(
        'account.fiscal.position', string='Fiscal Position',
        domain="[('company_id', '=', company_id)]", check_company=True,
        help="Fiscal positions are used to adapt taxes and accounts for particular customers or sales orders/invoices."
             "The default value comes from the customer.")
    company_id = fields.Many2one('res.company', 'Company', required=True, index=True,
                                 default=lambda self: self.env.company)
    team_id = fields.Many2one(
        'crm.team', 'Sales Team',
        change_default=True, default=_get_default_team, check_company=True,  # Unrequired company
        domain="['|', ('company_id', '=', False), ('company_id', '=', company_id)]")

    signature = fields.Image('Signature', help='Signature received through the portal.', copy=False, attachment=True,
                             max_width=1024, max_height=1024)
    signed_by = fields.Char('Signed By', help='Name of the person that signed the SO.', copy=False)
    signed_on = fields.Datetime('Signed On', help='Date of the signature.', copy=False)

    commitment_date = fields.Datetime('Delivery Date', copy=False,
                                      states={'done': [('readonly', True)], 'cancel': [('readonly', True)]},
                                      help="This is the delivery date promised to the customer. "
                                           "If set, the delivery order will be scheduled based on "
                                           "this date rather than product lead times.")
    amount_undiscounted = fields.Float('Amount Before Discount', compute='_compute_amount_undiscounted', digits=0)
    handling_info = fields.Text(string="Handling Information")
    accounting_info = fields.Text(string="Accounting Information")
    ref_num = fields.Char(string="Reference Number")
    optional_shipping = fields.Char(string="Optional Shipping Information")
    ppd = fields.Float(string="PPD")
    coll = fields.Float(string="COLL")
    other_ppd = fields.Float(string="PPD")
    other_coll = fields.Float(string="COLL")
    code = fields.Selection(
                selection = [
                    ('pp', 'PP'),
                    ('cc', 'CC'),
        ],     string = "CHGS Code")

    declare_value = fields.Char(string="Declared value for carriage")
    declare_value_custom = fields.Char(string="Declared value for custom")
    amount_insurance = fields.Float(string="Amount of Insurance")
    nature_goods = fields.Text(string="Nature and Quantity Of Goods")

    export = fields.Boolean('Exports')
    country_id = fields.Many2one('res.country', 'Country')
    port_of_loading_id = fields.Many2one('ports', 'Port of Loading')
    port_of_discharge_id = fields.Many2one('ports', 'Port of Discharge')
    cha_id = fields.Many2one('res.partner', 'Custom House Agent ')
    agent_code = fields.Char(string = 'Agent IATA Code')
    agent_account_number = fields.Char(string = 'Agent Account No.')
    freight_forwarder_id = fields.Many2one('res.partner', 'Freight Forwarder')
    stuffing_point_id = fields.Many2one('ports', 'Stuffing Point')
    cur_rate = fields.Float('Currency  Exchange Rate.')
    reference_by_id = fields.Many2one('res.partner', 'Reference By')
    # payment_term_ids = fields.Many2many('account.payment.term', 'payment_term_rel', 'order_id', 'payment_term_id',string='Payment Terms')
    sale_expense_line = fields.One2many('cargo.expense.line', 'order_id', 'Expenses')
    dry_port_id = fields.Many2one('ports', 'Dry/ICD Port')
    payment_term_line = fields.One2many('cargo.payment.term.line', 'order_id', 'Payment Details')
    sale_export_document_line = fields.One2many('cargo.export.document', 'order_id', 'Export Documents')
    total_expense = fields.Float(String="Total Expenses", store='True', compute='get_total_expense')
    rate_per_kg_on_expense = fields.Float(String="Rate Per Kg On Expense", store='True', compute='get_total_expense')
    is_load_ecgc_insurance = fields.Boolean(string="Is Load ECGC Insurance")
    schedule_lines = fields.One2many('cargo.schedule.line', 'sale_id', 'Schedule Lines')
    qc_pro_specification_line = fields.One2many('cargo.qc.specification.line', 'order_id',
                                                string='QC Product Specification Line')
    # team_id = fields.Many2one('crm.team', 'Sales Team', change_default=True, default=_get_default_team,
    #                           oldname='section_id')
    # user_id = fields.Many2one('res.users', string='Salesperson', index=True, track_visibility='onchange',
    #                           track_sequence=2, default=lambda self: self.env.user)
    export_tc_set_ids = fields.Many2many('term.and.condition.set', string="Terms & Condtions Set")
    export_so_tc_set_line = fields.One2many('cargo.so.tc.set.lines', 'export_so_tc_set_id',
                                            string='Terms And Condition Set Lines')
    ex_p_street = fields.Char('Street ', size=64)
    ex_p_street2 = fields.Char('Street2 ', size=64)
    ex_p_zip = fields.Char('Zip ', size=64)
    ex_p_city = fields.Char('City ', size=64)
    ex_p_state_id = fields.Many2one('res.country.state', string="State ")
    # origin1 = fields.Many2one('account.move', string="Origin")#by kajal
    ex_p_country_id = fields.Many2one('res.country', string="Country ")
    ex_d_street = fields.Char('Street', size=64)
    ex_d_street2 = fields.Char('Street2', size=64)
    ex_d_zip = fields.Char('Zip', size=64)
    ex_d_city = fields.Char('City', size=64)
    ex_d_state_id = fields.Many2one('res.country.state', string="State.")
    ex_d_country_id = fields.Many2one('res.country', string="Country.")
    consignee_id = fields.Many2one('res.partner', string="Consignee")
    consignee_account = fields.Char(string="Consignee Account No. ")
    consignor_id = fields.Many2one('res.partner', string="Consignor")
    consignor_account = fields.Char(string="Consignor Account No. ")
    notify_id = fields.Many2one('res.partner', string="Notify Party")
    so_invoice_id = fields.Many2one('so.invoice.value', string="So Inv Val")
    so_amount_total_usd = fields.Float(compute='_so_total_amount_usd', string='SO Amount In USD')
    packing_details = fields.Char(string="Packing Details")
    cargo_details_ids = fields.One2many('cargo.details.line', 'order_id', string="Cargo Details")
    # service_details_ids = fields.One2many('service.details.line', 'service_id', string="Service Details")

    # export_import = fields.Selection([('export', 'Export'), ('import', 'Import')], string='Export/Import')
    commodity = fields.Many2one('export.product.category', string="Commodity Type")
    hs_code = fields.Char(string='HS Code')
    shipment_by = fields.Selection([('sea', 'Sea'), ('air', 'Air')], string='Shipment By')
    shipping_line = fields.Many2one('res.partner', string="Shipping Line")
    air_airline_no = fields.Char(string='Airline No(Prefix)')
    air_airline_code = fields.Char(string='Airline Code')

    vessel_id = fields.Many2one('ocean.vessel', string="Ocean Vessel")
    voyage_id = fields.Many2one('ocean.voyage', string="Voyage")
    shipping_order_no = fields.Char(string="Shipping Order No.")
    carrier_booking_no = fields.Char(string="Carrier Booking No.")
    eta_port_of_loading = fields.Date(string="ETA - Port of Loading")
    cut_off_time = fields.Date(string="Cut off Time")
    estimated_time_departure = fields.Date(string="Estimated Time of Departure")
    actual_time_sailing = fields.Date(string="Actual Time of Sailing")
    eta_port_of_destination = fields.Date(string="ETA - Port of Destination")
    est_transit_days = fields.Date(string="Est. Transit Days ")
    bill_of_lading_type = fields.Many2one('bill.lading.type', string="Bill of Lading Type")
    bill_of_lading_no = fields.Char(string="Bill of Lading No.")
    # airway_master = fields.Many2one('awb.master', string="AWB Master")
    mawb = fields.Many2one('awb.master.line', string="MAWB")
    mawb_land = fields.Char(string="MAWB")
    hawb = fields.Char(string="HAWB")
    check_digit = fields.Integer("Check Digit")
    customer_po_ref = fields.Char(string="Customer’s PO Ref.")

    freight_forwarding = fields.Boolean(string="Freight Forwarding", default=False)
    pre_cargo_carriage = fields.Boolean(string="Pre Cargo Carriage", default=False)
    custom_clearance = fields.Boolean(string="Custom Clearance", default=False)
    reefer_dry = fields.Selection([
        ('reefer', 'Reefer'),
        ('dry', 'Dry')], string="Reefer/Dry")
    genset = fields.Boolean(string="Genset", default=False)
    gsa_sales = fields.Boolean(string="GSA Sales", default=False)

    # freight_forwarding_reqd = fields.Boolean(string="Freight Forwarding")
    # pre_cargo_carriage_reqd = fields.Boolean(string="Pre Cargo Carriage")
    # custom_clearance_reqd = fields.Boolean(string="Custom Clearance")
    # reefer_dry = fields.Selection([('reefer', 'Reefer'), ('dry', 'Dry')], string='Reefer/Dry')
    # genset_reqd = fields.Boolean(string="Genset")
    # gsa_sales = fields.Boolean(string="GSA Sales", default=False)
    purchase_id = fields.Many2one('purchase.order', string='Related Purchase Order')

    incoterm_id = fields.Many2one('account.incoterms', string="Shipment (Inco) Terms")
    service_type = fields.Many2one('service.type', string="Service Type")
    weight_uom_id = fields.Many2one('uom.uom', string='Weight UOM', default=_get_default_weight_uom_id)
    dimension_uom_id = fields.Many2one('uom.uom', string='Dimension UOM', default=_get_default_dimension_uom_id)
    cargo_freight_line = fields.One2many('cargo.freight.line', 'cargo_id', string='Cargo Freight Line', copy=True)
    total_pieces = fields.Float(string='Total Pieces', store=True, compute='_compute_sum_of_total')
    total_gross_weight = fields.Float(string='Total Gross Weight', store=True, compute='_compute_sum_of_total')
    total_cbm = fields.Float(string='Total CBM', store=True, compute='_compute_sum_of_total')
    total_value = fields.Float(string='Total Value', store=True, compute='_compute_sum_of_total')
    total_chargeable_weight = fields.Float(string='Total Chargeable weight', store=True,
                                           compute='_compute_total_chargeable_weight')
    total_volume_cbm = fields.Float(string='Total Volume (cbm)', store=True, compute='_compute_sum_of_total')
    total_volumetric_weight = fields.Float(string='Total Volumetric weight', store=True,
                                           compute='_compute_sum_of_total')
    pivot_weight = fields.Float(string='Pivot Weight', compute='_compute_sum_of_total', store=True)
    account_analytic_id = fields.Many2one('account.analytic.account', string='Analytic Account')
    opportunity_id = fields.Many2one('crm.lead', string='Opportunity ID')
    cargo_container_line = fields.One2many('cargo.container.lines', 'cargo_order_id', string='Container Line')
    filename = fields.Binary(string='Filename')
    track_trace_ids = fields.One2many('cargo.order.track.trace', 'cargo_order_id', 'Track & Trace')
    dept_date = fields.Date(string="Depart Date")

    total_sale_amount = fields.Monetary(string='Total Customer Order Amount', compute='_compute_amount_lead')
    total_po_amount = fields.Monetary(string='Total Vendor Order Amount', compute='_compute_amount_lead')
    total_invoice_amount = fields.Monetary(string='Total Invoice Amount', compute='_compute_amount_lead')
    total_bill_amount = fields.Monetary(string='Total Bill Amount', compute='_compute_amount_lead')
    profit_on_order = fields.Monetary(string="Profit Amount", compute='_compute_profit_amount')
    total_gross_weight_cargo = fields.Float(
        string="Total Gross Weight", compute="_compute_totals_cargo", store=True
    )
    total_chargeable_weight_cargo = fields.Float(
        string="Total Chargeable Weight", compute="_compute_totals_cargo", store=True
    )
    total_amount_cargo = fields.Float(
        string="Total Amount", compute="_compute_totals_cargo", store=True
    )
    cmc_cd = fields.Char(string="Cmc Cd")
    exchange_rate = fields.Char(string="Exchange Rate")
    # total_sales_amount = fields.Float(string="Total Sales Amount", compute="_compute_totals_service", store=True)
    # total_charge_total = fields.Float(string="Total Charge Total", compute="_compute_totals_service", store=True)
    customer_type = fields.Selection([
        ('individual', 'Individual'),
        ('corporate', 'Corporate')
    ], string='Cust. Type')
    mode_of_pay = fields.Selection([
        ('cash', 'Cash'),
        ('card', 'Card'),
        ('bank', 'Bank Transfer')
    ], string='Mode of Pay')

    purchase_amt = fields.Float(string='Purchase Amt')
    service_percent = fields.Float(string='Service %')
    service_amt = fields.Float(string='Service Amt')
    disc_percent = fields.Float(string='Disc %')
    disc_amt = fields.Float(string='Disc Amt')
    other_income_amt = fields.Float(string='Other Income Amt')
    sales_amt = fields.Float(string='Sales AMT')
    closing_detail_ids = fields.One2many(
        'closing.details.line',
        'closing_order_id',
        string='Closing Details'
    )
    total_purchase_charge = fields.Float(string='Total Purchase Charge')
    due_carrier = fields.Float(string='Due Carrier', compute='_compute_due_amounts', store=True)
    due_agent = fields.Float(string='Due Agent', compute='_compute_due_amounts', store=True)

    gsa_comm_perc = fields.Float(string='GSA Comm %')
    gsa_amount = fields.Float(string='GSA Amount')
    comm_perc = fields.Float(string='Comm %')
    comm_amount = fields.Float(string='Commission Amount')
    f_c_amt = fields.Float(string='F C Amt')
    amt_total = fields.Float(string='AMT')
    iata_rt = fields.Char("IATA RT")
    special_rt = fields.Char("Special RT")
    other_remarks = fields.Text("Other Remarks")
    agent_id = fields.Many2one('res.partner', string="Agent")
    payable_to = fields.Many2one('res.partner', string="Payable To")
    flight_date_1 = fields.Date(string="Flight Date 1")
    flight_number_1 = fields.Char(string="Flight Number 1")
    flight_date_2 = fields.Date(string="Flight Date 2")
    flight_number_2 = fields.Char(string="Flight Number 2")
    # First route
    to_1 = fields.Many2one('ports', string="To")
    to_2 = fields.Many2one('ports', string="To")
    to_3 = fields.Many2one('ports', string="To")

    by_1 = fields.Many2one('res.partner', string="By Last Carrier", domain="[('is_airline','=',True)]")
    by_2 = fields.Many2one('res.partner', string="By", domain="[('is_airline','=',True)]")
    by_3 = fields.Many2one('res.partner', string="By", domain="[('is_airline','=',True)]")

    airline_code_1 = fields.Char(string="Airline Code", compute="_compute_airline_codes", store=False)
    airline_code_2 = fields.Char(string="Airline Code", compute="_compute_airline_codes", store=False)
    airline_code_3 = fields.Char(string="Airline Code", compute="_compute_airline_codes", store=False)

    account_code = fields.Char(string='A/c Code')
    account_description = fields.Char(string='A/c Description')
    sub_code = fields.Char(string='Sub Code')
    trx_amt = fields.Float(string='Trx Amt')
    sign_of_shipper = fields.Char(string='Signature of Shipper or Agent')
    sign_of_carrier = fields.Char(string='Signature of Issuing Carrier or Agent')
    dimension_volume = fields.Selection([
                                        ('dimension', 'All Dimensions'),
                                        ('volume', 'Total Volume')
                                        ], string='Dimension/Volume')
    company_currency_id = fields.Many2one('res.currency', related='company_id.currency_id')
    total_cur_conversion_amount = fields.Monetary(
        string='Total After Currency Conversion',
        currency_field='company_currency_id',
        compute='_compute_total_cur_conversion_amount',
        store=True
    )
    total_after_currency_conversion = fields.Monetary(
        string='Total After Currency Conversion',
        currency_field='company_currency_id',
        compute='_compute_total_after_currency_conversion',
        store=True
    )
    move_id = fields.Many2one('account.move', string="Origin")#by kajal

    def unlink(self):
        for rec in self:
            # Check the linked invoice via move_id
            if rec.move_id:
                if rec.move_id.state == 'posted':
                    raise ValidationError(f"Cannot delete the Cargo Order: related invoice {rec.move_id.name} is posted.")
                elif rec.move_id.state == 'draft':
                    raise ValidationError(f"Cannot delete the Cargo Order: please cancel the draft invoice {rec.move_id.name} first.")
            # Reset MAWB state if used
            if rec.mawb and rec.mawb.state == 'used':
                rec.mawb.state = 'open'
        return super(CargoOrder, self).unlink()

    @api.depends('total_amount_cargo', 'cur_rate')
    def _compute_total_after_currency_conversion(self):
        for rec in self:
            rec.total_after_currency_conversion = rec.total_amount_cargo * rec.cur_rate if rec.cur_rate else 0.0

    @api.depends('amount_total', 'cur_rate')
    def _compute_total_cur_conversion_amount(self):
        for rec in self:
            rec.total_cur_conversion_amount = rec.amount_total * rec.cur_rate if rec.cur_rate else 0.0

    @api.depends('order_line.charge_type', 'order_line.price_total')
    def _compute_due_amounts(self):
        for record in self:
            agent_total = 0.0
            carrier_total = 0.0
            for line in record.order_line:
                if line.charge_type == 'due_agent':
                    agent_total += line.price_total
                elif line.charge_type == 'due_carrier':
                    carrier_total += line.price_total
            record.due_agent = agent_total
            record.due_carrier = carrier_total


    @api.depends('by_1.airline_code', 'by_2.airline_code', 'by_3.airline_code')
    def _compute_airline_codes(self):
        for rec in self:
            rec.airline_code_1 = rec.by_1.airline_code if rec.by_1 else ''
            rec.airline_code_2 = rec.by_2.airline_code if rec.by_2 else ''
            rec.airline_code_3 = rec.by_3.airline_code if rec.by_3 else ''

    # @api.depends('service_details_ids.sales_amount', 'service_details_ids.total_charge')
    # def _compute_totals_service(self):
    #     for record in self:
    #         record.total_sales_amount = sum(line.sales_amount for line in record.service_details_ids)
    #         record.total_charge_total = sum(line.total_charge for line in record.service_details_ids)

    @api.depends('cargo_details_ids', 'cargo_details_ids.gross_weight', 'cargo_details_ids.chargeable_weight', 'cargo_details_ids.total')
    def _compute_totals_cargo(self):
        for rec in self:
            rec.total_gross_weight_cargo = sum(line.gross_weight for line in rec.cargo_details_ids)
            rec.total_chargeable_weight_cargo = sum(line.chargeable_weight for line in rec.cargo_details_ids)
            rec.total_amount_cargo = sum(line.total for line in rec.cargo_details_ids)

    @api.onchange('shipping_line')
    def on_shipping_line(self):
        if self.shipping_line:
            self.air_airline_no = self.shipping_line.airline_no
            self.air_airline_code = self.shipping_line.airline_code
        else:
            self.air_airline_no = False
            self.air_airline_code = False

    # @api.onchange('shipping_line')
    # def _onchange_shipping_line(self):
    #     if self.shipping_line:
    #         return {
    #             'domain': {
    #                 'mawb': [('airline_id', '=', self.shipping_line.id), ('state', '=', 'open')]
    #             },
    #             'value': {
    #                 'mawb': False,
    #                 'check_digit': False
    #             }
    #         }

    @api.onchange('mawb')
    def _onchange_mawb(self):
        if self.mawb:
            self.check_digit = self.mawb.check_digit

    @api.depends('total_sale_amount', 'total_po_amount', 'total_expense')
    def _compute_profit_amount(self):
        for val in self:
            val.profit_on_order = val.total_sale_amount - val.total_po_amount - val.total_expense

    def action_view_agent_orders_purchase(self):
        self.ensure_one()
        result = {
            "type": "ir.actions.act_window",
            "res_model": "purchase.order",
            "domain": [('opportunity_id', '=',  self.opportunity_id.id), ('state', "=", 'purchase')],
            "context": {"create": False},
            "name": "Vendor Orders",
            'view_mode': 'tree,form',
        }
        return result

    def action_view_customer_orders_sale(self):
        self.ensure_one()
        result = {
            "type": "ir.actions.act_window",
            "res_model": "sale.order",
            "domain": [('opportunity_id', '=', self.opportunity_id.id), ('state', "=", 'sale')],
            "context": {"create": False},
            "name": "Customer orders",
            'view_mode': 'tree,form',
        }
        return result

    @api.depends('opportunity_id')
    def _compute_amount_lead(self):
        for order in self:
            total_amount_so = total_amount_po = total_amount_inv = total_amount_bill = 0.0
            if order.opportunity_id:
                sale_orders = self.env['sale.order'].search([('opportunity_id', '=', self.opportunity_id.id), ('state', '=', 'sale')])
                # total_amount_so = sum(order.amount_total for order in sale_orders)

                if sale_orders:
                    for so in sale_orders:
                        invs = so.invoice_ids.filtered(lambda inv: inv.state == 'posted' and inv.payment_state == 'not_paid')
                        total_amount_inv = sum(i.amount_total_signed for i in invs)

                        if so.currency_id == so.company_currency_id:
                            total_amount_so += so.amount_total
                        else:
                            total_amount_so += so.total_cur_conversion_amount

                purchase_orders = self.env['purchase.order'].search([('opportunity_id', '=', self.opportunity_id.id), ('state', '=', 'purchase')])
                # total_amount_po = sum(order.amount_total for order in purchase_orders)
                # for po in purchase_orders:
                #     if order.currency_id == order.company_currency_id:
                #         total_amount_po += order.amount_total
                #     else:
                #         total_amount_po += order.total_cur_conversion_amount
                if purchase_orders:
                    for po in purchase_orders:
                        invs = po.invoice_ids.filtered(lambda inv: inv.state == 'posted' and inv.payment_state == 'not_paid')
                        total_amount_bill = sum(i.amount_total_signed for i in invs)

                        if po.currency_id == po.company_currency_id:
                            total_amount_po += po.amount_total
                        else:
                            total_amount_po += po.total_cur_conversion_amount


            order.total_sale_amount = total_amount_so
            order.total_po_amount = total_amount_po
            order.total_invoice_amount = total_amount_inv
            order.total_bill_amount = total_amount_bill

    @api.onchange('commodity')
    def _onchange_commodity(self):
        if self.commodity:
            self.hs_code = self.commodity.hs_code
        else:
            self.hs_code = False

    def upload_container_deatils(self):

        for self_obj in self:
            data_decode = self_obj.filename
            if not data_decode:
                raise UserError(_('Please Choose The File!'))
            val = base64.decodebytes(data_decode)
            fp = BytesIO()
            fp.write(val)
            wb = xlrd.open_workbook(file_contents=fp.getvalue())
            sheet_name = wb.sheet_names()
            sh = wb.sheet_by_name(sheet_name[0])
            n_rows = sh.nrows
            for row in range(1, n_rows):
                if sh.row_values(row)[0]:
                    container_id = self.env['container.type'].search([('name', '=', sh.row_values(row)[0])], limit=1)
                    if not container_id:
                        container_id = self.env['container.type'].create({
                            "name": sh.row_values(row)[0]
                        })
                    cargo_container_lines = self.env['cargo.container.lines'].create({
                        'container_type_id': container_id.id,
                        'count': sh.row_values(row)[1],
                        'container_qty': sh.row_values(row)[2],
                        'cargo_order_id': self_obj.id
                    })

    @api.depends('cargo_freight_line', 'cargo_freight_line.product_uom_quantity', 'cargo_freight_line.volume',
                 'cargo_freight_line.total_weight', 'cargo_freight_line.commercial_invoice_value',
                 'cargo_freight_line.total_chargeable_weight', 'cargo_freight_line.total_volume_cbm',
                 'cargo_freight_line.total_volumetric_weight', 'cargo_freight_line.pivot_weight')
    def _compute_sum_of_total(self):
        for res in self:
            total_pieces = 0.0
            total_gross_weight = 0.0
            total_cbm = 0.0
            total_value = 0.0
            total_chargeable_weight = 0.0
            total_volume_cbm = 0.0
            total_volumetric_weight = 0.0
            total_pivot_weight = 0.0
            for line in res.cargo_freight_line:
                total_pieces += line.product_uom_quantity
                total_gross_weight += line.total_weight
                total_cbm += line.volume
                total_value += line.commercial_invoice_value
                total_chargeable_weight += line.total_chargeable_weight
                total_volume_cbm += line.total_volume_cbm
                total_volumetric_weight += line.total_volumetric_weight
                total_pivot_weight += line.pivot_weight
            res.total_pieces = total_pieces
            res.total_gross_weight = total_gross_weight
            res.total_cbm = total_cbm
            res.total_value = total_value
            # res.total_chargeable_weight = total_chargeable_weight
            res.total_volume_cbm = total_volume_cbm
            res.total_volumetric_weight = total_volumetric_weight
            res.pivot_weight = total_pivot_weight

    @api.depends('total_gross_weight', 'total_volumetric_weight', 'pivot_weight')
    def _compute_total_chargeable_weight(self):
        for rec in self:
            rec.total_chargeable_weight = max(
                rec.total_gross_weight or 0,
                rec.total_volumetric_weight or 0,
                rec.pivot_weight or 0
            )

    @api.model
    def create(self, vals):
        """Cargo Order Create Function"""
        if vals.get('name', _('New')) == _('New') and vals.get('order_type') == 'freight' \
                and vals.get('mode') == 'sea' and vals.get('import_export') == 'import':
            print("if freight/sea/import====================")
            vals['name'] = self.env['ir.sequence'].next_by_code('cargo.sea.import.seq')
        elif vals.get('name', _('New')) == _('New') and vals.get('order_type') == 'freight' \
                and vals.get('mode') == 'sea' and vals.get('import_export') == 'export':
            print("if freight/sea/export====================")
            vals['name'] = self.env['ir.sequence'].next_by_code('cargo.sea.export.seq')
        elif vals.get('name', _('New')) == _('New') and vals.get('order_type') == 'freight' \
                and vals.get('mode') == 'air' and vals.get('import_export') == 'import':
            print("if freight/air/import====================")
            vals['name'] = self.env['ir.sequence'].next_by_code('cargo.air.import.seq')
        elif vals.get('name', _('New')) == _('New') and vals.get('order_type') == 'freight' \
                and vals.get('mode') == 'air' and vals.get('import_export') == 'export':
            print("if freight/air/export====================")
            vals['name'] = self.env['ir.sequence'].next_by_code('cargo.air.export.seq')
        elif vals.get('name', _('New')) == _('New') and vals.get('order_type') == 'custom' \
                and vals.get('import_export') == 'import':
            print("if custom/import====================")
            vals['name'] = self.env['ir.sequence'].next_by_code('cargo.clearance.import.seq')
        elif vals.get('name', _('New')) == _('New') and vals.get('order_type') == 'custom' \
                and vals.get('import_export') == 'export':
            print("if custom/export====================")
            vals['name'] = self.env['ir.sequence'].next_by_code('cargo.clearance.export.seq')
        elif vals.get('name', _('New')) == _('New') and vals.get('order_type') == 'freight' \
                and vals.get('mode') == 'land' and vals.get('import_export') == 'import':
            print("if freight/land/import====================")
            vals['name'] = self.env['ir.sequence'].next_by_code('cargo.land.import.seq')
        elif vals.get('name', _('New')) == _('New') and vals.get('order_type') == 'freight' \
                and vals.get('mode') == 'land' and vals.get('import_export') == 'export':
            print("if freight/land/export====================")
            vals['name'] = self.env['ir.sequence'].next_by_code('cargo.land.export.seq')

        elif vals.get('name', _('New')) == _('New') and vals.get('mode') == 'courier' \
                and vals.get('import_export') == 'import':
            print("if courier/import====================")
            vals['name'] = self.env['ir.sequence'].next_by_code('cargo.courier.import.seq')

        elif vals.get('name', _('New')) == _('New') and vals.get('mode') == 'courier' \
                and vals.get('import_export') == 'export':
            print("if coruier/export====================")
            vals['name'] = self.env['ir.sequence'].next_by_code('cargo.courier.export.seq')

        elif vals.get('name', _('New')) == _('New') and vals.get('mode') == 'air' \
                and vals.get('import_export') == 'cross_trade':
            print("if air/cross_trade====================")
            vals['name'] = self.env['ir.sequence'].next_by_code('cargo.air.cross_trade.seq')

        elif vals.get('name', _('New')) == _('New') and vals.get('mode') == 'sea' \
                and vals.get('import_export') == 'cross_trade':
            print("if sea/cross_trade====================")
            vals['name'] = self.env['ir.sequence'].next_by_code('cargo.sea.cross_trade.seq')

        # elif vals.get('name', _('New')) == _('New') and vals.get('order_type') == 'packing_and_removal':
        #     vals['name'] = self.env['ir.sequence'].next_by_code('packing.removal.cargo.seq')
        # elif vals.get('name', _('New')) == _('New') and vals.get('order_type') == 'warehousing':
        #     vals['name'] = self.env['ir.sequence'].next_by_code('warehousing.cargo.seq')
        else:
            print("else====================")
            vals['name'] = self.env['ir.sequence'].next_by_code('cargo.order.seq')
        if self.env.context and self.env.context.get('default_order_type') == 'freight':
            vals['state'] = 'cargo_ready'
        if self.env.context and self.env.context.get('default_order_type') == 'custom':
            vals['state'] = 'prealert_received'
        if vals.get('order_type') == 'freight':
            vals['state'] = 'cargo_ready'
        if vals.get('order_type') == 'custom':
            vals['state'] = 'prealert_received'
        res = super(CargoOrder, self).create(vals)
        return res

    # def _compute_origin_value(self):
    #     if self.partner_id:
    #         self.name = self.origin1.co_id.id
    #
    #     else:
    #         self.name = self.origin1.co_id.id

    # def _compute_profatibility(self):
    #     bill = self.origin1.search([('co_id','=',self.id),('move_type','=','in_invoice')])
    #     invoice = self.origin1.search([('co_id','=',self.id),('move_type','=','out_invoice')])
    #     if bill and invoice:
    #         final_amt=bill.amount_untaxed-invoice.amount_untaxed
    #         self.profitability = final_amt
    #     else:
    #         self.profitability =0.00 ##by kajal

    # def action_get_bill(self):
    #     itemIds = self.env['account.move'].search([])
    #     itemIds = itemIds.ids
    #     return {
    #         'name': _('Vendor Bill'),
    #         'type': 'ir.actions.act_window',
    #         'view_type': 'form',
    #         'view_mode': 'tree,form',
    #         'res_model': 'account.move',
    #         'view_id': False,
    #         'context': {'default_move_type': 'in_invoice','default_co_id':self.id},
    #         'domain': [('id', 'in', itemIds),('move_type','=','in_invoice')],
    #     }  ##by kajal

    # def action_get_invoice(self):
    #     itemIds = self.env['account.move'].search([])
    #     itemIds = itemIds.ids
    #     return {
    #         'name': ('Customer Invoice'),
    #         'type': 'ir.actions.act_window',
    #         'view_type': 'form',
    #         'view_mode': 'tree,form',
    #         'res_model': 'account.move',
    #         'view_id': False,
    #         'context':{'default_move_type': 'out_invoice','default_co_id':self.id},
    #         'domain': [('id', 'in', itemIds),('move_type','=','out_invoice')],
    #     }##by kajal

    # @api.model
    # def create(self, vals):
    #     '''Cargo Order Create Function'''
    #     if vals.get('name', _('New')) == _('New'):
    #         vals['name'] = self.env['ir.sequence'].next_by_code('cargo.order.seq')
    #     if vals.get('internal_type') in ('sea_outbound', 'sea_inbound', 'air_outbound', 'air_inbound'):
    #         vals['state'] = 'cargo_ready'
    #     if vals.get('internal_type') == 'custom_clearance':
    #         vals['state'] = 'prealert_received'
    #     res = super(CargoOrder, self).create(vals)
    #     return res

    def write(self, vals):
        '''Cargo Order Write Function'''
        # if vals.get('internal_type') in ('sea_outbound', 'sea_inbound', 'air_outbound', 'air_inbound'):
        #     vals['state'] = 'cargo_ready'
        # if vals.get('internal_type') == 'custom_clearance':
        #     vals['state'] = 'prealert_received'
        if vals.get('order_type') == 'freight':
            vals['state'] = 'cargo_ready'
        if vals.get('order_type') == 'custom':
            vals['state'] = 'prealert_received'
        if 'state' in vals:
            field_list = self._fields['state'].selection
            if type(field_list) == list:
                selection_dict = dict(field_list)
                self.env['cargo.order.track.trace'].create({
                    'date': fields.date.today(),
                    'user_id': self.env.user.id,
                    'cargo_order_id': self.id,
                    'status': selection_dict.get(vals['state'])
                })
        res = super(CargoOrder, self).write(vals)
        return res

    @api.onchange('partner_id')
    def onchange_partner_id(self):
        if self.partner_id:
            self.pricelist_id = self.partner_id.property_product_pricelist.id

    def action_cro_requested(self):
        self.state = 'cro_requested'

    def action_cro_released(self):
        self.state = 'cro_released'

    def action_stuff_started(self):
        self.state = 'stuff_started'

    def action_stuffing_completed(self):
        self.state = 'stuffing_completed'

    def action_sent_to_port(self):
        self.state = 'sent_to_port'

    def action_custom_clearance(self):
        self.state = 'custom_clearance'

    def action_sealing_done(self):
        self.state = 'sealing_done'

    def action_bill_of_lading_completed(self):
        self.state = 'bill_of_lading_completed'

    def action_vessel_set_to_sail(self):
        self.state = 'vessel_set_to_sail'

    def action_pre_alert(self):
        self.state = 'pre_alert'

    def action_destination_arrival(self):
        self.state = 'destination_arrival'

    def action_destination_clearance(self):
        self.state = 'destination_clearance'

    def action_delivery(self):
        self.state = 'delivery'

    def action_completed(self):
        for rec in self:
            if rec.purchase_id:
                if rec.purchase_id.state not in ['purchase', 'done']:
                    raise UserError(
                        f"Please confirm the vendor order first. Related Vendor Order: {rec.purchase_id.name}"
                    )
            if rec.mode == 'air':
                if rec.mawb and rec.mawb.state == 'used':
                    raise UserError("This MAWB no is already in used state.")
                if rec.mawb:
                    rec.mawb.state = 'used'
            rec.state = 'completed'
   
    def action_set_to_draft(self):
        if self.move_id:
            raise UserError(
                f"Set to Draft is not possible because invoice is already created."
            )
        else:
            if self.order_type == 'freight':
                self.state = 'cargo_ready'
            if self.order_type == 'custom':
                self.state = 'prealert_received' 

    
    def action_book_with_flight(self):
        self.state = 'book_with_the_flight'

    def action_arrange_the_airway_bill(self):
        self.state = 'arrange_the_airway_bill'

    def action_pickup(self):
        self.state = 'pickup'

    def action_sent_to_port(self):
        self.state = 'sent_to_port'

    def action_booking_flight_schedule_confirmation(self):
        self.state = 'booking_flight_schedule_confirmation'

    def action_eta_shipment_arrival(self):
        self.state = 'eta_shipment_arrival'

    def action_document_readiness(self):
        self.state = 'document_readiness'

    def action_cargo_arrival_notice(self):
        self.state = 'cargo_arrival_notice'

    def action_forward_for_clearance(self):
        self.state = 'forward_for_clearance'

    def action_under_clearance(self):
        self.state = 'under_clearance'

    def action_awaiting_approval(self):
        self.state = 'awaiting_approval'

    def action_custom_declaration_completed(self):
        self.state = 'custom_declaration_completed'

    def action_under_custom_inspection(self):
        self.state = 'under_custom_inspection'

    def action_out_for_delivery(self):
        self.state = 'out_for_delivery'

    def action_send_to_finance(self):
        name_list = []
        for res in self:
            if not res.sale_id or not res.sale_id.order_line:
                raise ValidationError(_('No Lines To Send Finance!'))
            if not res.partner_id:
                raise ValidationError(_(
                    'You can not make invoice without partner. Please select customer in case there is no customer.'))

            if res.order_type:
                journal_id = self.env['account.journal'].sudo().search(
                    [('type', '=', 'sale'),
                     ('order_type', '=', res.order_type),
                     ('mode', '=', res.mode),
                     ('import_export', '=', res.import_export)], limit=1)
                if not journal_id:
                    journal_id = self.env['account.journal'].sudo().search([('type', '=', 'sale')], limit=1)
            else:
                journal_id = self.env['account.journal'].sudo().search([('type', '=', 'sale')], limit=1)

            total_pieces = 0.0
            for cl in res.cargo_details_ids:
                total_pieces += cl.no_of_pieces

            invoice_vals = {
                'partner_id': res.partner_id.id or False,
                'move_type': 'out_invoice',
                'invoice_date': datetime.today(),
                'journal_id': journal_id.id or False,
                'invoice_line_ids': [],
                'is_export': True,
                'ref': res.name,
                'consignee_id': res.consignee_id.id or False,
                'consignor_id': res.consignor_id.id or False,
                'notify_id': res.notify_id.id or False,
                'cha_id': res.cha_id.id or False,
                'reference_by_id': res.reference_by_id.id or False,
                'cur_rate': res.sale_id.cur_rate or res.cur_rate,
                'stuffing_point_id': res.stuffing_point_id.id or False,
                'port_of_loading_id': res.port_of_loading_id.id or False,
                'port_of_discharge_id': res.port_of_discharge_id.id or False,
                'order_type': res.order_type,
                'mode': res.mode,
                'import_export': res.import_export,
                'invoice_incoterm_id': res.incoterm_id.id or False,
                'account_analytic_id': res.account_analytic_id.id or False,
                'customer_po_ref': res.customer_po_ref,
                'bill_of_lading_no': res.bill_of_lading_no,
                'mawb': res.mawb.id or False,
                'hawb': res.hawb,
                'mawb_land': res.mawb_land,
                'estimated_time_departure': res.estimated_time_departure,
                'eta_port_of_destination': res.eta_port_of_destination,
                'total_pieces': res.total_pieces,
                'total_gross_weight': res.total_gross_weight if res.mode != 'air' else res.total_gross_weight_cargo,
                'total_chargeable_weight': res.total_chargeable_weight if res.mode != 'air' else res.total_chargeable_weight_cargo,
                'total_cbm': res.total_cbm,
                'total_value': res.total_value,
                'total_volume_cbm': res.total_volume_cbm,
                'total_volumetric_weight': res.total_volumetric_weight,
                'ref_num': res.ref_num,
                'date_order': res.date_order,
                'flight_date_1': res.flight_date_1,
                'flight_number_1': res.flight_number_1,
                'cargo_id': res.id,
                'shipping_line': res.shipping_line.id or False,
                'air_airline_code': res.air_airline_code,
                'air_airline_no': res.air_airline_no,
                'currency_id': res.currency_id.id or False,
                'freight_forwarding': res.freight_forwarding,
                'pre_cargo_carriage': res.pre_cargo_carriage,
                'custom_clearance': res.custom_clearance,
                'reefer_dry': res.reefer_dry,
                'genset': res.genset,
                'gsa_sales': res.gsa_sales,
            }

            # ✅ Properly link SO lines so invoice_status is updated
            for lines in res.sale_id.order_line:
                dict1 = {
                    'name': lines.name,
                    'product_uom_id': lines.product_uom.id,
                    'product_id': lines.product_id.id,
                    'price_unit': lines.price_unit,
                    'quantity': lines.product_uom_qty,
                    'tax_ids': [(6, 0, lines.tax_id.ids)],
                    'sale_line_ids': [(6, 0, [lines.id])],
                }
                invoice_vals['invoice_line_ids'].append((0, 0, dict1))

            # Expense lines (not linked to SO lines, so don’t affect invoice_status)
            for exp_line in res.sale_expense_line:
                dict2 = {
                    'name': exp_line.exp_related_to,
                    'product_id': exp_line.expense_id.id,
                    'product_uom_id': exp_line.expense_id.uom_id.id,
                    'price_unit': exp_line.rate,
                    'quantity': exp_line.qty,
                }
                invoice_vals['invoice_line_ids'].append((0, 0, dict2))

            # Create invoice
            inv_id = self.env['account.move'].with_context(
                manual_currency_rate=invoice_vals.get('cur_rate')
            ).create(invoice_vals)

            if inv_id:
                res.move_id = inv_id.id
                for lines in res.cargo_container_line:
                    vals = {
                        'move_id': inv_id.id,
                        'container_type_id': lines.container_type_id.id or False,
                        'count': lines.count,
                        'container_qty': lines.container_qty,
                    }
                    self.env['move.container.lines'].create(vals)

            return {
                'type': 'ir.actions.act_window',
                'view_mode': 'form',
                'res_model': 'account.move',
                'target': 'current',
                'res_id': inv_id.id,
            }


class CargoOrderLine(models.Model):
    _name = 'cargo.order.line'
    _description = 'Cargo Order Line'
    _order = 'order_id, sequence, id'
    _check_company_auto = True

    @api.depends('product_uom_qty', 'discount', 'price_unit', 'tax_id')
    def _compute_amount(self):
        """
        Compute the amounts of the SO line.
        """
        for line in self:
            price = line.price_unit * (1 - (line.discount or 0.0) / 100.0)
            taxes = line.tax_id.compute_all(price, line.order_id.currency_id, line.product_uom_qty,
                                            product=line.product_id, partner=line.order_id.partner_shipping_id)
            line.update({
                'price_tax': sum(t.get('amount', 0.0) for t in taxes.get('taxes', [])),
                'price_total': taxes['total_included'],
                'price_subtotal': taxes['total_excluded'],
            })
            if self.env.context.get('import_file', False) and not self.env.user.user_has_groups(
                    'account.group_account_manager'):
                line.tax_id.invalidate_cache(['invoice_repartition_line_ids'], [line.tax_id.id])

    order_id = fields.Many2one('cargo.order', string='Order Reference', required=True, ondelete='cascade', index=True,
                               copy=False)
    name = fields.Text(string='Description', required=True)
    gross_weight = fields.Float(string="Gross Weight")
    rate_class = fields.Char(string="Rate Class")
    chargeable_weight = fields.Float(string="Chargeable Weight")
    remark = fields.Text(string='Remark')
    sequence = fields.Integer(string='Sequence', default=10)
    price_unit = fields.Float('Unit Price', required=True, digits='Product Price', default=0.0)
    price_subtotal = fields.Monetary(compute='_compute_amount', string='Subtotal', readonly=True, store=True)
    price_tax = fields.Float(compute='_compute_amount', string='Total Tax', readonly=True, store=True)
    price_total = fields.Monetary(compute='_compute_amount', string='Total', readonly=True, store=True)
    tax_id = fields.Many2many('account.tax', string='Taxes',
                              domain=['|', ('active', '=', False), ('active', '=', True)])
    discount = fields.Float(string='Discount (%)', digits='Discount', default=0.0)
    product_id = fields.Many2one(
        'product.product', string='Product',
        domain="[('sale_ok', '=', True), '|', ('company_id', '=', False), ('company_id', '=', company_id)]",
        change_default=True, ondelete='restrict', check_company=True)  # Unrequired company
    product_template_id = fields.Many2one(
        'product.template', string='Product Template',
        related="product_id.product_tmpl_id", domain=[('sale_ok', '=', True)])
    product_uom_qty = fields.Float(string='Quantity', digits='Product Unit of Measure', required=True, default=1.0)
    product_uom = fields.Many2one('uom.uom', string='Unit of Measure',
                                  domain="[('category_id', '=', product_uom_category_id)]")
    product_uom_category_id = fields.Many2one(related='product_id.uom_id.category_id', readonly=True)
    product_uom_readonly = fields.Boolean(compute='_compute_product_uom_readonly')

    untaxed_amount_invoiced = fields.Monetary("Untaxed Invoiced Amount", compute='_compute_untaxed_amount_invoiced',
                                              compute_sudo=True, store=True)
    untaxed_amount_to_invoice = fields.Monetary("Untaxed Amount To Invoice",
                                                compute='_compute_untaxed_amount_to_invoice', compute_sudo=True,
                                                store=True)

    salesman_id = fields.Many2one(related='order_id.user_id', store=True, string='Salesperson', readonly=True)
    currency_id = fields.Many2one(related='order_id.currency_id', depends=['order_id.currency_id'], store=True,
                                  string='Currency', readonly=True)
    company_id = fields.Many2one(related='order_id.company_id', string='Company', store=True, readonly=True, index=True)
    order_partner_id = fields.Many2one(related='order_id.partner_id', store=True, string='Customer', readonly=False)
    is_expense = fields.Boolean('Is expense',
                                help="Is true if the sales order line comes from an expense or a vendor bills")
    state = fields.Selection(
        related='order_id.state', string='Order Status', readonly=True, copy=False, store=True)

    display_type = fields.Selection([
        ('line_section', "Section"),
        ('line_note', "Note")], default=False, help="Technical field for UX purpose.")

    base_price = fields.Float(string="Base Price", digits=('16', 2))
    overhead_expenses = fields.Float(string="Overhead/Expenses")
    packing_type = fields.Many2one('export.packing.type', string="Packing Type")
    bag_box_qty = fields.Float(string="Bag/Box Qty.")
    container_type_id = fields.Many2one('container.type', 'Container Type')
    capacity_in_mt = fields.Float('Capacity(MT)')
    show_in_awb = fields.Boolean('Show in AWB', defalt=False)
    charge_type = fields.Selection([
                                    ('due_agent', "Due Agent"),
                                    ('due_carrier', "Due Carrier")], string="Charge Type")

    def _compute_tax_id(self):
        for line in self:
            line = line.with_company(line.company_id)
            fpos = line.order_id.fiscal_position_id or line.order_id.fiscal_position_id.get_fiscal_position(
                line.order_partner_id.id)
            # If company_id is set, always filter taxes by the company
            taxes = line.product_id.taxes_id.filtered(lambda t: t.company_id == line.env.company)
            line.tax_id = fpos.map_tax(taxes, line.product_id, line.order_id.partner_shipping_id)

    @api.onchange('container_type_id')
    def onchange_container_type_id(self):
        if self.container_type_id:
            self.capacity_in_mt = self.container_type_id.capacity

    @api.onchange('product_id')
    def product_id_change(self):
        if not self.product_id:
            return
        if self.product_id:
            self.product_uom = self.product_id.uom_id
            self.product_uom_qty = self.product_uom_qty or 1.0
            self.name = self.product_id.display_name
            self.tax_id = [(6, False, self.product_id.taxes_id.ids)]

    @api.onchange('product_id', 'price_unit', 'product_uom', 'product_uom_qty', 'tax_id')
    def _onchange_discount(self):
        if not (self.product_id and self.product_uom and
                self.order_id.partner_id and self.order_id.pricelist_id and
                self.order_id.pricelist_id.discount_policy == 'without_discount' and
                self.env.user.has_group('product.group_discount_per_so_line')):
            return

        self.discount = 0.0
        product = self.product_id.with_context(
            lang=self.order_id.partner_id.lang,
            partner=self.order_id.partner_id,
            quantity=self.product_uom_qty,
            date=self.order_id.date_order,
            pricelist=self.order_id.pricelist_id.id,
            uom=self.product_uom.id,
            fiscal_position=self.env.context.get('fiscal_position')
        )

        product_context = dict(self.env.context, partner_id=self.order_id.partner_id.id, date=self.order_id.date_order,
                               uom=self.product_uom.id)

        price, rule_id = self.order_id.pricelist_id.with_context(product_context).get_product_price_rule(
            self.product_id, self.product_uom_qty or 1.0, self.order_id.partner_id)
        new_list_price, currency = self.with_context(product_context)._get_real_price_currency(product, rule_id,
                                                                                               self.product_uom_qty,
                                                                                               self.product_uom,
                                                                                               self.order_id.pricelist_id.id)

        if new_list_price != 0:
            if self.order_id.pricelist_id.currency_id != currency:
                # we need new_list_price in the same currency as price, which is in the SO's pricelist's currency
                new_list_price = currency._convert(
                    new_list_price, self.order_id.pricelist_id.currency_id,
                    self.order_id.company_id or self.env.company, self.order_id.date_order or fields.Date.today())
            discount = (new_list_price - price) / new_list_price * 100
            if (discount > 0 and new_list_price > 0) or (discount < 0 and new_list_price < 0):
                self.discount = discount


class CargoFreightline(models.Model):
    _name = "cargo.freight.line"
    _description = "Cargo Freight Line"

    def _get_default_dimension_uom_id(self):
        return self.env.ref('uom.product_uom_cm')

    def _get_default_weight_uom_id(self):
        return self.env.ref('uom.product_uom_kgm')

    cargo_id = fields.Many2one('cargo.order', string="Cargo Id")
    product_id = fields.Many2one('product.product', string="Product")
    container_type_id = fields.Many2one('container.type', string="Container Type")
    capacity_in_mt = fields.Float(string="Capacity (MT)")
    container_qty = fields.Integer(string="Container Qty.")
    commercial_invoice_value = fields.Float(string="Commercial Invoice Value")
    goods_desc = fields.Text(string='Goods Description')
    packing_type = fields.Many2one('export.packing.type', string="Packing Type")
    product_uom = fields.Many2one('uom.uom', string="Unit of Measurement")
    # no_of_package = fields.Integer(string='No. Of Package')
    type = fields.Selection([
        ('stackable', 'Stackable'),
        ('non_stackable', 'Non-Stackable')], string='Type', help="Are packages stackable?")
    product_uom_quantity = fields.Float(string="No. of Bag/Box/Pack")
    weight_input = fields.Float(string='Weight')
    weight_per_package = fields.Float(string='Weight Per Package')
    total_weight = fields.Float(string='Total Weight', store=True, compute='_compute_total_weight')
    loose_bool = fields.Boolean(string="Loose Quantity", default=False)
    length_per_package = fields.Float(string='length Per Package(cm)')
    width_per_package = fields.Float(string='Width Per Package(cm)')
    height_per_package = fields.Float(string='Height Per Package(cm)')
    weight_factor = fields.Float(string='Weight Factor')
    chargeable_weight = fields.Float(string='Chargeable weight Per pack', store=True,
                                     compute='_compute_chargeable_weight')
    total_chargeable_weight = fields.Float(string='Total Chargeable weight', store=True,
                                           compute='_compute_total_chargeable_weight')
    volume_cbcm = fields.Float(string='Volume (Cubic cm)', store=True, compute='_compute_volume_cbcm')
    volume = fields.Float(string='Volume Per Package(cbm)', store=True, compute='_compute_volume_cbm')
    volumetric_weight = fields.Float(string='Volumetric Weight', store=True, compute='_compute_volumetric_weight',
                                     help="If weight factor type in service type is by Cubic CM then volumetric weight = volume/weight factor else volume*weight factor.")
    weight_uom_id = fields.Many2one('uom.uom', string='Weight UOM', default=_get_default_weight_uom_id)
    dimension_uom_id = fields.Many2one('uom.uom', string='Dimension UOM', default=_get_default_dimension_uom_id)
    uom_name = fields.Char(string="UoM Name", related='dimension_uom_id.name')
    remarks = fields.Text(string="Remarks")
    full_container_service = fields.Boolean(string="Full Container Service", default=False)
    # subtotal = fields.Float(string='Subtotal', compute='_compute_subtotal')
    total_volume_cbm = fields.Float(string='Total Volume (cbm)', store=True, compute='_compute_total_volume_cbm')
    total_volumetric_weight = fields.Float(string='Total Volumetric weight', store=True,
                                           compute='_compute_total_volumetric_weight')
    pivot_weight = fields.Float(string='Pivot Weight')
    hs_code = fields.Char('HS code')
    commodity_type = fields.Many2one('export.product.category', string="Commodity Type")

    @api.onchange('product_id')
    def onchange_product_id(self):
        if self.product_id:
            self.product_uom = self.product_id.uom_id.id

    @api.onchange('container_type_id')
    def onchange_container_type_id(self):
        if self.container_type_id:
            self.capacity_in_mt = self.container_type_id.capacity

    @api.onchange('product_id', 'cargo_id.service_type')
    def onchange_get_weight_factor(self):
        if self.cargo_id.service_type:
            self.weight_factor = self.cargo_id.service_type.weight_factor
            self.full_container_service = self.cargo_id.service_type.full_container_service

    @api.onchange('product_id', 'cargo_id.weight_uom_id', 'cargo_id.dimension_uom_id')
    def onchange_get_uoms(self):
        if self.cargo_id.weight_uom_id:
            self.weight_uom_id = self.cargo_id.weight_uom_id.id
        if self.cargo_id.dimension_uom_id:
            self.dimension_uom_id = self.cargo_id.dimension_uom_id.id

    @api.depends('weight_per_package', 'product_uom_quantity')
    def _compute_total_weight(self):
        for res in self:
            res.total_weight = res.weight_per_package * res.product_uom_quantity

    # @api.depends('total_weight', 'volume')
    # def _compute_subtotal(self):
    #     for res in self:
    #         res.subtotal = res.price_unit * (res.total_weight if res.total_weight > res.volume else res.volume)

    @api.depends('length_per_package', 'width_per_package', 'height_per_package')
    def _compute_volume_cbcm(self):
        for res in self:
            res.volume_cbcm = res.length_per_package * res.width_per_package * res.height_per_package

    # @api.depends('length_per_package', 'width_per_package', 'height_per_package')
    # def _compute_volume_cbm(self):
    #     for res in self:
    #         res.volume = (res.length_per_package * res.width_per_package * res.height_per_package) / 1000000

    @api.depends('volume_cbcm', 'dimension_uom_id', 'length_per_package', 'width_per_package', 'height_per_package')
    def _compute_volume_cbm(self):
        for res in self:
            if res.dimension_uom_id.name == 'm':
                res.volume = (res.length_per_package * res.width_per_package * res.height_per_package)
            elif res.dimension_uom_id.name == 'cm':
                res.volume = res.volume_cbcm / 1000000
            elif res.dimension_uom_id.name == 'in':
                res.volume = res.volume_cbcm / 61024
            else:
                res.volume = 0

    @api.depends('chargeable_weight', 'product_uom_quantity')
    def _compute_total_chargeable_weight(self):
        for res in self:
            res.total_chargeable_weight = res.chargeable_weight * res.product_uom_quantity

    @api.depends('volume_cbcm', 'weight_factor', 'cargo_id.service_type')
    def _compute_volumetric_weight(self):
        for res in self:
            if res.weight_factor > 0:
                if res.cargo_id.service_type.weight_factor_type == 'on_cbcm':
                    res.volumetric_weight = res.volume_cbcm / res.weight_factor
                else:
                    res.volumetric_weight = res.volume * res.weight_factor

    @api.onchange('weight_input', 'weight_per_package')
    def onchange_weight_input(self):
        # for res in self:
        if self.weight_uom_id.name == 'kg':
            self.weight_per_package = self.weight_input
        elif self.weight_uom_id.name == 'lb':
            self.weight_per_package = self.weight_input / 2.205
        else:
            self.weight_per_package = 0

    @api.depends('weight_per_package', 'volumetric_weight')
    def _compute_chargeable_weight(self):
        for res in self:
            if res.weight_per_package > res.volumetric_weight:
                res.chargeable_weight = res.weight_per_package
            else:
                res.chargeable_weight = res.volumetric_weight

    @api.depends('volume', 'product_uom_quantity')
    def _compute_total_volume_cbm(self):
        for res in self:
            res.total_volume_cbm = res.product_uom_quantity * res.volume

    @api.depends('volumetric_weight', 'product_uom_quantity')
    def _compute_total_volumetric_weight(self):
        for res in self:
            res.total_volumetric_weight = res.product_uom_quantity * res.volumetric_weight


class CargoExpenseLine(models.Model):
    _name = "cargo.expense.line"
    _description = 'Cargo Expense Line'

    @api.depends('qty', 'rate')
    def _get_exp_amt(self):
        for val in self:
            val.exp_amt = val.qty * val.rate

    order_id = fields.Many2one('cargo.order', 'Order Id')
    exp_related_to = fields.Char('Expense Related to')
    # product_class = fields.Many2one('product.classification', string='Product Class')
    expense_id = fields.Many2one('product.product', 'Expense Name')
    qty = fields.Integer(string='Quantity', default=1)
    usd_value = fields.Float(string='USD')
    rate = fields.Float('Rate')
    exp_amt_fc = fields.Float('Expense Amt (FC)')
    exp_amt = fields.Float('Expense Amt', compute='_get_exp_amt')
    remark = fields.Char('Remarks')

    # invoice_id = fields.Many2one('account.invoice', string="Invoice")

    # @api.onchange('expense_id')
    # def onchange_expense_id(self):
    #     if self.expense_id:
    #         self.product_class = self.expense_id.product_class.id

    @api.onchange('usd_value')
    def onchange_usd_value(self):
        for res in self:
            if res.usd_value > 0:
                res.rate = res.usd_value * res.order_id.cur_rate

    @api.onchange('rate')
    def onchange_rate(self):
        for res in self:
            if res.rate > 0 and res.order_id.cur_rate > 0:
                res.usd_value = res.rate / res.order_id.cur_rate


class CargoExportDocument(models.Model):
    _name = "cargo.export.document"
    _description = 'Cargo Export Document'

    export_document_id = fields.Many2one('export.document', 'Document Name')
    order_id = fields.Many2one('cargo.order', 'Cargo Order Id')
    # invoice_id = fields.Many2one('account.invoice', string="Invoice") ##by kajal


class CargoPaymenTermLine(models.Model):
    _name = "cargo.payment.term.line"
    _description = 'Cargo Payment term Line'

    order_id = fields.Many2one('cargo.order', 'Order Id')
    country_id = fields.Many2one('res.country', 'Country')
    payment_id = fields.Many2one('account.payment.term', 'Payment Term')
    value = fields.Float(string='Value (%)')
    ecgc_rate = fields.Float(string='ECGC Rate (%)')

    # invoice_id = fields.Many2one('account.invoice', string="Invoice") ##by kajal

    @api.onchange('payment_id')
    def get_ecgc_rate(self):
        for res in self:
            if self._context.get('country_id'):
                country_id = self.env["res.country"].browse(self._context.get('country_id'))
                if country_id:
                    ecgc_id = self.env["ecgc.rating.line"].search([('country_id', '=', country_id.id)], limit=1)
                    payment_id = self.env["ecgc.premium.rate"].search(
                        [('ecgc_rating_id', '=', ecgc_id.ecgc_rating_id.id)])
                    for payment in payment_id:
                        if payment.payment_term_id == res.payment_id:
                            res.ecgc_rate = payment.rate

    @api.onchange('payment_id')
    def onchange_payment_id(self):
        payment_ids_list = []
        if self._context.get('country_id'):
            country_id = self.env["res.country"].browse(self._context.get('country_id'))
            if country_id:
                ecgc_id = self.env["ecgc.rating.line"].search([('country_id', '=', country_id.id)], limit=1)
                payment_id = self.env["ecgc.premium.rate"].search([('ecgc_rating_id', '=', ecgc_id.ecgc_rating_id.id)])
                for payment in payment_id:
                    payment_ids_list.append(payment.payment_term_id.id)
        domain = {'payment_id': [('id', 'in', payment_ids_list)]}
        return {'domain': domain}


class CargoDetailsLine(models.Model):
    _name = 'cargo.details.line'
    _description = 'Cargo Details Line'

    order_id = fields.Many2one('cargo.order', string="Cargo Order", ondelete="cascade")
    no_of_pieces = fields.Float(string="No. of Pcs / RCP")
    remarks = fields.Char(string="Commodity")
    class_cargo = fields.Char(string="Rate Class")
    gross_weight = fields.Float(string="Gross Weight")
    uom_id = fields.Many2one('uom.uom', string="UoM")
    chargeable_weight = fields.Float(string="Chargeable Weight")
    rate = fields.Float(string="Rate / Charge")
    total = fields.Float(string="Total", compute="_compute_total")

    @api.depends('chargeable_weight', 'rate')
    def _compute_total(self):
        for line in self:
            line.total = line.chargeable_weight * line.rate

# class ServiceDetailsLine(models.Model):
#     _name = 'service.details.line'
#     _description = 'Service Details Line'
#
#     service_id = fields.Many2one('cargo.order', string="Cargo Order", ondelete="cascade")
#     cost_code = fields.Char(string='Cost Code')
#     purchase_rate = fields.Float(string='Purchase rate')
#     total_charge = fields.Float(string='Total Charge')
#     sales_rate = fields.Float(string='Sales rate')
#     sales_amount = fields.Float(string='Sales Amount')




class ClosingDetailLine(models.Model):
    _name = 'closing.details.line'
    _description = 'Closing Detail Line'

    closing_order_id = fields.Many2one('cargo.order', string='Cargo Order')

    sales_type = fields.Selection([
        ('open', 'Open'),
        ('closed', 'Closed')
    ], string='Sales Type')

    sales_date = fields.Date(string='Sales Date')
    doc_no = fields.Char(string='Doc#')
    sales_amt = fields.Float(string='Sales Amt')
    disc_pct = fields.Float(string='Disc Pct')
    disc_amt = fields.Float(string='Disc Amt')
    rcvd_amt = fields.Float(string='Rcvd Amt')
    revenue = fields.Float(string='Revenue')
    incentive = fields.Float(string='Incentive')

    # account_code = fields.Char(string='A/c Code')
    # account_description = fields.Char(string='A/c Description')
    # sub_code = fields.Char(string='Sub Code')
    # trx_amt = fields.Float(string='Trx Amt')


class CargoScheduleLine(models.Model):
    _name = "cargo.schedule.line"
    _description = "Cargo Schedule Line"
    _rec_name = "schedule_id"

    sale_id = fields.Many2one('cargo.order', string="Order Id")
    schedule_id = fields.Many2one('common.schedule', string="Schedule Name")
    report_at_time = fields.Selection([('pre_shipment', 'Pre Shipment'), ('post_shipment', 'Post Shipment')],
                                      string='Report at Time')
    req_from = fields.Selection([('other', 'Other'), ('self', 'Self')], string='Requirement From')
    country_specific = fields.Boolean(string="Country Specific")
    port_of_loading_specific = fields.Boolean(string="Port of Loading Specific")
    completion_date = fields.Date(string="Completion Date")
    fname = fields.Char()
    attachment = fields.Binary(string="Attachment")
    remarks = fields.Text(string="Remarks")
    # invoice_id = fields.Many2one('account.invoice', string="Invoice")##by kajal
    #     export_do_id = fields.Many2one('stock.picking', string="DO")
    sale_export_attachment_lines = fields.One2many('cargo.schedule.attachment.line', 'sale_export_attachment_id',
                                                   string='Attachment Lines')

    @api.onchange('schedule_id')
    def onchange_schedule_id(self):
        if self.schedule_id:
            self.report_at_time = self.schedule_id.report_at_time
            self.req_from = self.schedule_id.req_from
            self.country_specific = self.schedule_id.country_specific
            self.port_of_loading_specific = self.schedule_id.port_of_loading_specific


class CargoScheduleAttachmentLine(models.Model):
    """ Attachment Details """
    _name = "cargo.schedule.attachment.line"
    _description = 'Cargo Schedule Attachment Line'

    sale_export_attachment_id = fields.Many2one('cargo.schedule.line')
    sale_remark = fields.Char(string='Remark')
    fname = fields.Char()
    sale_attachment_value = fields.Binary(string='Attachment')


class CargoQcSpecificationLine(models.Model):
    _name = "cargo.qc.specification.line"
    _description = "Cargo Qc Specification Line"
    _rec_name = 'qc_parameter_id'

    order_id = fields.Many2one('cargo.order', string="Order ID")
    product_id = fields.Many2one('product.product', string="Product")
    qc_parameter_id = fields.Many2one('product.mix.variants', string="Parameter")
    value = fields.Char(string="Value")
    picking_id = fields.Many2one('stock.picking', string="Picking")

    # invoice_id = fields.Many2one('account.invoice', string="Invoice")##by kajal

    @api.onchange('qc_parameter_id')
    def onchange_qc_parameter_id(self):
        if self.qc_parameter_id:
            self.value = self.qc_parameter_id.value


class CargoSoTcSetLines(models.Model):
    _name = "cargo.so.tc.set.lines"
    _description = "Cargo So Tc Set Lines"

    s_no = fields.Integer(string="S No", compute="_sequence_ref")
    export_so_tc_set_id = fields.Many2one('cargo.order', string='Terms And Condition Set Id')
    term_and_condition_id = fields.Many2one('term.and.condition', string='Condition')
    removable = fields.Selection([('yes', 'Yes'), ('no', 'No')],
                                 help='If user is allowed to remove this condition in the form such as SO/PO etc then choose it as Yes.',
                                 string='Removable')
    term_and_condition_option_id = fields.Many2one('term.and.condition.option', string='Terms And Condition Option')

    @api.depends('export_so_tc_set_id.export_so_tc_set_line')
    def _sequence_ref(self):
        for line in self:
            no = 0
            for l in line.export_so_tc_set_id.export_so_tc_set_line:
                no += 1
                l.s_no = no


class Cargocontainerlines(models.Model):
    _name = "cargo.container.lines"
    _description = "Cargo Container Line"

    cargo_order_id = fields.Many2one('cargo.order', string="Cargo Order ID")
    container_type_id = fields.Many2one('container.type', string="Container Type")
    count = fields.Integer(string="Count")
    container_qty = fields.Integer(string="Container Numbers")


class CargoOrderTrackTrace(models.Model):
    _name = "cargo.order.track.trace"

    cargo_order_id = fields.Many2one('cargo.order', string="Cargo Order ID")
    date = fields.Date("Date")
    user_id = fields.Many2one('res.users', string="User")
    status = fields.Char("Status")
